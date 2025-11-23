import pytest
import os
import json
import pytest
from openpyxl import Workbook, load_workbook

from app import create_app
from app.utils import get_db_connection
from app.routes import rates as rates_module


@pytest.fixture
def client(tmp_path, monkeypatch):
    """
    Create Flask test client with:
    - Clean DB before and after each test
    - Temporary /in folder for rates files (instead of /app/in)
    """

    # Create Flask app
    app = create_app()
    app.config["TESTING"] = True

    # Create a temporary folder that will act as /app/in
    in_dir = tmp_path / "in"
    in_dir.mkdir()

    # Monkeypatch os.path.join *inside* rates.py so that:
    # os.path.join("/app", "in", filename) → <temp_folder>/filename
    def fake_join(base, sub, filename):
        # We ignore base + sub ("/app", "in") and always join to in_dir
        return str(in_dir / filename)

    monkeypatch.setattr(rates_module.os.path, "join", fake_join)

    with app.test_client() as test_client:
        # Attach the temp rates folder to the client so tests can use it
        test_client.in_dir = in_dir

        # Clean DB before each test
        with app.app_context():
            cleanup_database()

        yield test_client

        # Clean DB after each test
        with app.app_context():
            cleanup_database()


def cleanup_database():
    """Delete all data from related tables before/after tests."""
    conn = get_db_connection()
    cursor = conn.cursor()

    # Clean in dependency-safe order
    cursor.execute("DELETE FROM Trucks")
    cursor.execute("DELETE FROM Rates")
    cursor.execute("DELETE FROM Provider")

    # Reset Provider AUTOINCREMENT like in provider tests
    cursor.execute("ALTER TABLE Provider AUTO_INCREMENT = 10001")

    conn.commit()
    cursor.close()
    conn.close()


def create_rates_excel(path, rows):
    """
    Helper: create an Excel file with columns:
    Product | Rate | Scope

    rows: list of tuples/lists → (product, rate, scope)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"

    # Header row
    ws.append(["Product", "Rate", "Scope"])

    # Data rows
    for product, rate, scope in rows:
        ws.append([product, rate, scope])

    wb.save(path)


# =======================
# Tests for POST /rates
# =======================

def test_post_rates_success(client):
    """
    POST /rates with a valid Excel file should:
    - Parse the file
    - Replace all rows in Rates table
    - Return status 200 with 'status': 'ok' and 'count'
    """
    # 1) Prepare Excel file in the temp rates folder
    excel_path = client.in_dir / "new_rates.xlsx"
    create_rates_excel(
        excel_path,
        rows=[
            ("1", 50, "ALL"),
            ("1", 70, "10001"),
            ("2", 40, "ALL"),
        ],
    )

    # 2) Call POST /rates with JSON body { "file": "new_rates.xlsx" }
    response = client.post(
        "/rates",
        data=json.dumps({"file": "new_rates.xlsx"}),
        content_type="application/json",
    )

    assert response.status_code == 200
    data = json.loads(response.data)
    assert data.get("status") == "ok"
    assert data.get("count") == 3

    # 3) Verify DB actually contains the inserted rows
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT product_id, rate, scope FROM Rates ORDER BY product_id, scope")
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    assert len(rows) == 3
    assert rows[0]["product_id"] == "1"
    assert rows[0]["rate"] == 50
    assert rows[0]["scope"].upper() == "ALL"


def test_post_rates_missing_file_param(client):
    """
    POST /rates without 'file' field should return 400.
    """
    response = client.post(
        "/rates",
        data=json.dumps({}),
        content_type="application/json",
    )

    assert response.status_code == 400
    data = json.loads(response.data)
    assert "error" in data


def test_post_rates_file_not_found(client):
    """
    POST /rates with a non-existing filename should return 404.
    """
    response = client.post(
        "/rates",
        data=json.dumps({"file": "no_such_file.xlsx"}),
        content_type="application/json",
    )

    assert response.status_code == 404
    data = json.loads(response.data)
    assert "File no_such_file.xlsx not found" in data.get("error", "")


# ======================
# Tests for GET /rates
# ======================

def test_get_rates_empty_db(client):
    """
    GET /rates with empty Rates table should return 404
    with a clear error message.
    """
    # DB was already cleaned in fixture
    response = client.get("/rates")

    assert response.status_code == 404
    data = json.loads(response.data)
    assert data.get("error") == "No rates found in the database."


def test_get_rates_generates_excel_and_download(client):
    """
    GET /rates with existing rows in DB should:
    - Generate an Excel file in the temp /in folder
    - Return HTTP 200
    - Return an XLSX with correct headers
    """

    # 1) Insert some rates directly to DB
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO Rates (product_id, rate, scope) VALUES (%s, %s, %s)",
        ("1", 50, "ALL"),
    )
    cursor.execute(
        "INSERT INTO Rates (product_id, rate, scope) VALUES (%s, %s, %s)",
        ("2", 40, "10001"),
    )
    conn.commit()
    cursor.close()
    conn.close()

    # 2) Call GET /rates
    response = client.get("/rates")

    assert response.status_code == 200
    # Check content type is Excel
    assert (
        response.content_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # 3) Verify that the file was created in the temp /in directory
    excel_path = client.in_dir / "rates.xlsx"
    assert excel_path.exists()

    # 4) Read the Excel and verify content
    wb = load_workbook(excel_path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    # rows[0] is header, rows[1:] are data
    assert rows[0] == ("product_id", "rate", "scope")
    assert len(rows) == 1 + 2  # header + 2 rows
